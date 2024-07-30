import pandas as pd
from openpyxl import load_workbook
import shutil

# Rutas a los archivos
template_path = 'files/prueba/parte2/FDE_Ventas_2023_3.xlsm'  # Ruta al archivo de plantilla original
ventas_path = 'files/prueba/parte2/ventas.xlsx'  # Ruta al archivo de ventas
output_path = 'files/prueba/parte2/FDE_Ventas_2023_8.xlsm'  # Ruta al archivo de salida

# Hacer una copia de la plantilla para trabajar sobre ella
shutil.copyfile(template_path, output_path)

# Leer las hojas en DataFrames
fde_ventas_df = pd.read_excel(template_path, sheet_name='FDE', engine='openpyxl')
ventas_df = pd.read_excel(ventas_path)

# Limpiar y preparar los DataFrames para la combinación
fde_ventas_clean_df = fde_ventas_df.iloc[4:].reset_index(drop=True)  # Eliminar las primeras 4 filas y reiniciar el índice
fde_ventas_clean_df.columns = fde_ventas_df.iloc[3]  # Usar la fila 4 como encabezados de columna
fde_ventas_clean_df = fde_ventas_clean_df.rename(columns={'Ref. Cliente': 'Ref.Cliente'})  # Renombrar la columna 'Ref. Cliente' a 'Ref.Cliente'

# Renombrar las columnas en el DataFrame de ventas
ventas_clean_df = ventas_df.rename(columns={
    'Ref. Cliente': 'Ref.Cliente', 
    'Canal_Zona': 'Canal_Zona', 
    'EXPORT': 'EXPORT', 
    'NACIONAL': 'NACIONAL',
    'Unnamed: 7': 'imp'  # Renombrar la columna 'Unnamed: 7' a 'imp'
})

# Asegurar que 'Ref.Cliente' sea tratado como cadena en ambos DataFrames
fde_ventas_clean_df['Ref.Cliente'] = fde_ventas_clean_df['Ref.Cliente'].astype(str)
ventas_clean_df['Ref.Cliente'] = ventas_clean_df['Ref.Cliente'].astype(str)

# Combinar los DataFrames en 'Ref.Cliente'
merged_df = pd.merge(fde_ventas_clean_df, ventas_clean_df[['Ref.Cliente', 'Canal_Zona', 'EXPORT', 'NACIONAL', 'imp']], on='Ref.Cliente', how='left')

# Multiplicar los valores por las unidades en la columna Q 'Nº de UD'
merged_df['Nº de UD'] = merged_df['Nº de UD'].astype(float)  # Asegurar que sea float para la multiplicación

# Definir las funciones condicionales para calcular los valores
def calculate_import(row):
    # Si 'imp' está en la columna 'imp', calcular 'Import' como 'NACIONAL' * 'Nº de UD'
    if 'imp' in str(row['imp']).lower():
        return row['NACIONAL'] * row['Nº de UD']
    return 0

def calculate_no_total(row):
    # Si 'imp' está en la columna 'imp', y 'EXPORT' tiene un valor, calcular 'Nº total' como 'EXPORT' * 'Nº de UD'
    if 'imp' in str(row['imp']).lower():
        if pd.notna(row['EXPORT']):
            return row['EXPORT'] * row['Nº de UD']
        return 0
    # Si 'imp' no está en la columna 'imp', calcular 'Nº total' como 'Canal_Zona' * 'Nº de UD'
    return row['Canal_Zona'] * row['Nº de UD'] if pd.notna(row['Canal_Zona']) else 0

def calculate_export(row):
    # Si 'imp' está en la columna 'imp', calcular 'Export' como 'EXPORT' * 'Nº de UD'
    if 'imp' in str(row['imp']).lower():
        return row['EXPORT'] * row['Nº de UD'] if pd.notna(row['EXPORT']) else 0
    # Si 'imp' no está en la columna 'imp', calcular 'Export' como 'EXPORT' * 'Nº de UD'
    return row['EXPORT'] * row['Nº de UD'] if pd.notna(row['EXPORT']) else 0

# Aplicar las funciones condicionales
merged_df['Nº total'] = merged_df.apply(calculate_no_total, axis=1)
merged_df['Export'] = merged_df.apply(calculate_export, axis=1)
merged_df['Import'] = merged_df.apply(calculate_import, axis=1)

# Actualizar los valores en el DataFrame original
fde_ventas_clean_df['Nº total'] = merged_df['Nº total']
fde_ventas_clean_df['Import'] = merged_df['Import']
fde_ventas_clean_df['Export'] = merged_df['Export']

# Comprobación de los datos antes de escribir
print(fde_ventas_clean_df[['Nº total', 'Import', 'Export']].head())

# Cargar el libro original y actualizar la hoja 'FDE'
book = load_workbook(output_path, keep_vba=True)
writer = pd.ExcelWriter(output_path, engine='openpyxl')

# Cargar la hoja y actualizar las celdas correspondientes
sheet = book['FDE']
for idx, row in fde_ventas_clean_df.iterrows():
    sheet.cell(row=6 + idx, column=22, value=row['Nº total'])  # Columna V (22)
    sheet.cell(row=6 + idx, column=23, value=row['Import'])    # Columna W (23)
    sheet.cell(row=6 + idx, column=24, value=row['Export'])    # Columna X (24)

# Guardar el libro actualizado
book.save(output_path)

print("Archivo actualizado guardado en 'FDE_Ventas_2023_8.xlsm'")
